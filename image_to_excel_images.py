import cv2
import numpy as np
import os
import re
import openpyxl
import pytesseract

# ==============================================================================
# CONFIGURATION
# ==============================================================================
IMAGE_FOLDER      = r"C:\Users\Tim\Desktop\PDF_IMAGE"
OUTPUT_IMAGES_DIR = r"C:\Users\Tim\Desktop\PDF_IMAGE_SAMPLE\IMAGES"
OUTPUT_EXCEL_PATH = r"C:\Users\Tim\Desktop\PDF_IMAGE_SAMPLE\new.xlsx"
DEBUG_FOLDER      = r"C:\Users\Tim\Desktop\PDF_IMAGE_SAMPLE\catalogue_debug"

DEBUG = True

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

PART_PATTERN = re.compile(r'^[A-Z]{2,4}\d{3,6}[A-Z]?(-\d+)?$')

# Original working picture crop coordinates (do not change these)
PICTURE_X0_PCT = 0.08
PICTURE_X1_PCT = 0.21

os.makedirs(OUTPUT_IMAGES_DIR, exist_ok=True)
if DEBUG:
    os.makedirs(DEBUG_FOLDER, exist_ok=True)

supported_exts = (".png", ".jpg", ".jpeg", ".bmp")
image_files = sorted([f for f in os.listdir(IMAGE_FOLDER) if f.lower().endswith(supported_exts)])
print(f"Found {len(image_files)} page images in {IMAGE_FOLDER}")


# ==============================================================================
# HELPERS
# ==============================================================================

def ocr_cell(cell_img, single_line=False):
    if cell_img is None or cell_img.size == 0:
        return ""
    h, w = cell_img.shape[:2]
    up = cv2.resize(cell_img, (w * 2, h * 2), interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(up, cv2.COLOR_BGR2GRAY)
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    cfg = '--psm 7 --oem 3' if single_line else '--psm 6 --oem 3'
    return pytesseract.image_to_string(binary, config=cfg).strip()


def strip_artifacts(text):
    """Remove leading OCR border artifacts (any non-alphanumeric char) from text fields."""
    return re.sub(r'^[^\w]+', '', text).strip()


def clean_oem(text):
    """Clean OEM reference: strip border artifacts and fix O→0 (OEM refs are purely numeric)."""
    text = strip_artifacts(text)
    # OEM numbers never contain the letter O — replace all O with 0
    text = text.replace('O', '0').replace('o', '0')
    return text


def clean_part_no(raw):
    """Fix common OCR misreads in part numbers before pattern matching."""
    s = re.sub(r'\s+', '', raw).upper()
    if len(s) < 5:
        return s
    # O and Z never appear as letters in this catalogue's part numbers — replace globally first
    s = s.replace('O', '0').replace('Z', '2')
    # Now find the true letter prefix (only real alphabet chars remain)
    letters = re.match(r'^[A-Z]+', s)
    if not letters:
        return s
    prefix = letters.group()
    remainder = s[len(prefix):]
    # Split off any hyphen-variant suffix (e.g. -1, -2) before cleaning digits
    _hyp = remainder.split('-', 1)
    digits = _hyp[0]
    variant = ('-' + _hyp[1]) if len(_hyp) > 1 else ''
    # Fix remaining digit lookalikes in the numeric suffix
    digits = digits.replace('I', '1').replace('L', '1')
    digits = digits.replace('S', '5').replace('B', '8').replace('G', '6')
    # OCR often reads a single 0 as two chars (0 + O → 00 after O→0 replacement).
    # e.g. AH001901 → AH01901, AH002004 → AH02004
    if digits.startswith('00') and len(digits) >= 4:
        digits = digits[1:]
    return prefix + digits + variant


def get_row_y(img):
    """Detect horizontal table row lines using Canny edges."""
    _, W = img.shape[:2]
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    edges = cv2.Canny(gray, 30, 100)
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (int(W * 0.35), 1))
    h_lines = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, h_kernel)
    cnts = cv2.findContours(h_lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[0]
    ys = sorted([cv2.boundingRect(c)[1] for c in cnts if cv2.boundingRect(c)[2] > W * 0.3])
    # Fallback to adaptive threshold
    if len(ys) < 3:
        thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                       cv2.THRESH_BINARY_INV, 51, 15)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (int(W * 0.35), 1))
        lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel, iterations=2)
        cnts2 = cv2.findContours(lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[0]
        ys = sorted([cv2.boundingRect(c)[1] for c in cnts2 if cv2.boundingRect(c)[2] > W * 0.3])
    clean = []
    for y in ys:
        if not clean or y - clean[-1] > 15:
            clean.append(y)
    H = img.shape[0]
    if not clean or clean[0] > 30:
        clean = [0] + clean
    if clean[-1] < H - 30:
        clean = clean + [H]
    return clean


def get_col_x(W):
    """
    Fixed column boundaries anchored to the original working picture crop (8%–21%).
    Auto-detection removed: part images create false vertical edges that shift all indices.

    Col: 0=S/N | 1=Picture | 2=Our No. | 3=Description | 4=OEM | 5=Application | 6=Unit
    """
    return [0, int(W*0.08), int(W*0.21), int(W*0.30), int(W*0.44), int(W*0.55), int(W*0.92), W]


def save_debug_img(img, row_y, col_x, filename):
    vis = img.copy()
    H, W = vis.shape[:2]
    for y in row_y:
        cv2.line(vis, (0, y), (W, y), (0, 200, 0), 2)
    for x in col_x:
        cv2.line(vis, (x, 0), (x, H), (0, 0, 220), 2)
    # Draw picture crop boundaries in orange
    cv2.line(vis, (int(W * PICTURE_X0_PCT), 0), (int(W * PICTURE_X0_PCT), H), (0, 165, 255), 2)
    cv2.line(vis, (int(W * PICTURE_X1_PCT), 0), (int(W * PICTURE_X1_PCT), H), (0, 165, 255), 2)
    cv2.imwrite(os.path.join(DEBUG_FOLDER, f"debug_{filename}"), vis)


# ==============================================================================
# MAIN PROCESSING
# ==============================================================================

all_extracted_rows = []

for page_num, filename in enumerate(image_files):
    img_path = os.path.join(IMAGE_FOLDER, filename)
    img = cv2.imread(img_path)
    if img is None:
        print(f"  ⚠ Could not read {filename}")
        continue

    H, W = img.shape[:2]
    row_y = get_row_y(img)
    col_x = get_col_x(W)

    print(f"\n{'='*60}")
    print(f"{filename}  [{W}×{H}px]")
    print(f"  Row lines ({len(row_y)}): {row_y[:10]}{'...' if len(row_y)>10 else ''}")
    print(f"  Col x     : {col_x}")

    if DEBUG:
        save_debug_img(img, row_y, col_x, filename)

    for i in range(len(row_y) - 1):
        y0, y1 = row_y[i], row_y[i + 1]
        if y1 - y0 < 40:
            continue

        def crop(col_idx):
            if col_idx + 1 >= len(col_x):
                return None
            return img[y0:y1, col_x[col_idx]:col_x[col_idx + 1]]

        # OCR the Our No. cell — single-line mode, then fix common misreads
        our_no_raw = ocr_cell(crop(2), single_line=True)
        our_no = clean_part_no(our_no_raw)

        matched = bool(PART_PATTERN.match(our_no))

        if DEBUG:
            print(f"  row {i:2d} y={y0}-{y1}  raw='{our_no_raw}'  clean='{our_no}'  match={matched}")
            # Save crop for ALL rows so we can diagnose skipped ones
            cell = crop(2)
            if cell is not None and cell.size > 0:
                label = our_no if matched else f"SKIP_{i}_{our_no}"
                cv2.imwrite(os.path.join(DEBUG_FOLDER, f"cell_{label}.png"), cell)

        if not matched:
            continue

        _desc_raw = ocr_cell(crop(3)).replace('\n', ' ').strip()

        # Check raw text first: if it starts with -N the column cut the variant suffix
        # e.g. our_no="AH06301"  raw="-1 Piston..."  →  "AH06301-1" / "Piston..."
        _var = re.match(r'^-(\d+)\s*(.*)', _desc_raw, re.DOTALL)
        if _var:
            our_no = our_no + '-' + _var.group(1)
            description = strip_artifacts(_var.group(2).strip())
        else:
            description = strip_artifacts(_desc_raw)
            # If description starts with a single digit + space, the OCR column boundary
            # cut the last digit of the part number into the description column — move it back.
            # e.g.  our_no="ABP3201"  description="8 Brake Pad"  →  "ABP32018" / "Brake Pad"
            _dm = re.match(r'^(\d)\s+(.*)', description, re.DOTALL)
            if _dm:
                _candidate = our_no + _dm.group(1)
                if PART_PATTERN.match(_candidate):
                    our_no = _candidate
                    description = strip_artifacts(_dm.group(2).strip())

        oem         = clean_oem(ocr_cell(crop(4)).replace('\n', ' '))
        application = strip_artifacts(ocr_cell(crop(5)).replace('\n', ' '))
        unit        = strip_artifacts(ocr_cell(crop(6)).replace('\n', ' ')) or "PC"

        # Picture crop: use original working coordinates (8%–21%)
        pic = img[y0:y1, int(W * PICTURE_X0_PCT):int(W * PICTURE_X1_PCT)]

        image_filename = f"{our_no}.png"
        if pic is not None and pic.size > 0:
            cv2.imwrite(os.path.join(OUTPUT_IMAGES_DIR, image_filename), pic)

        all_extracted_rows.append({
            "Source Page":  filename,
            "Our No.":      our_no,
            "Description":  description,
            "OEM":          oem,
            "Application":  application,
            "Unit":         unit,
            "Image File":   image_filename,
        })
        print(f"  ✓ {our_no:12s} | {description[:45]}")

# ==============================================================================
# EXCEL GENERATION
# ==============================================================================

if not all_extracted_rows:
    print("\n⚠ No rows matched PART_PATTERN.")
    print("  Check the 'raw=' debug lines above — if text is garbled/empty,")
    print("  the column boundaries or OCR are off.")
    print(f"  Debug images → {DEBUG_FOLDER}")
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parts Catalogue"
    ws.views.sheetView[0].showGridLines = True

    headers = ["Source Page", "Our No.", "Description", "OEM", "Application", "Unit", "Image File"]
    ws.append(headers)

    header_font  = openpyxl.styles.Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill  = openpyxl.styles.PatternFill(start_color="2B4C7E", end_color="2B4C7E", fill_type="solid")
    center_align = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    left_wrap    = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border  = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin', color='D1D5DB'),
        right=openpyxl.styles.Side(style='thin', color='D1D5DB'),
        top=openpyxl.styles.Side(style='thin', color='D1D5DB'),
        bottom=openpyxl.styles.Side(style='thin', color='D1D5DB'),
    )
    CENTER_COLS = {1, 2, 6, 7}

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    ws.row_dimensions[1].height = 28

    for r_idx, row_data in enumerate(all_extracted_rows, start=2):
        vals = [row_data["Source Page"], row_data["Our No."], row_data["Description"],
                row_data["OEM"], row_data["Application"], row_data["Unit"], row_data["Image File"]]
        ws.append(vals)
        ws.row_dimensions[r_idx].height = 50
        for c_idx in range(1, len(vals) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = openpyxl.styles.Font(name="Segoe UI", size=10)
            cell.border = thin_border
            cell.alignment = center_align if c_idx in CENTER_COLS else left_wrap

    for col in ws.columns:
        max_len = max((len(str(c.value).split('\n')[0]) for c in col if c.value), default=10)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 55)

    wb.save(OUTPUT_EXCEL_PATH)
    print(f"\n✅ Done — {len(all_extracted_rows)} parts extracted.")
    print(f"   Excel  → {OUTPUT_EXCEL_PATH}")
    print(f"   Images → {OUTPUT_IMAGES_DIR}")
