#!/usr/bin/env python3
"""
Import docs/speedgrand.xlsx → Supabase

  Sheet1       → parts
  SupplierItem → part_suppliers  (auto-creates missing suppliers)

Column mapping:
  ItemId       → sku
  Description  → name
  NewCost      → cost_price
  PriceAVat    → price
  Bin          → bin_location
  Catalogory   → category
  Make         → make
  Model        → model
  Year         → year_range

Usage:
  python import_speedgrand.py [--dry-run] [--skip-parts] [--skip-suppliers]

  --dry-run        Print counts only, do not write to Supabase
  --skip-parts     Skip Sheet1 import (parts already loaded)
  --skip-suppliers Skip SupplierItem import
"""

import openpyxl
import urllib.request
import urllib.parse
import urllib.error
import json
import sys
import os

# Force UTF-8 output on Windows
if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SUPABASE_URL = "https://lskouiyvdngdzaquurhk.supabase.co"
SUPABASE_KEY = "sb_publishable_De4neqOoFn1wFyiVzaNT0A_HzPAE3YW"
XLSX_PATH    = os.path.join(os.path.dirname(__file__), "docs", "speedgrand.xlsx")
BATCH_SIZE   = 500

DRY_RUN       = "--dry-run"       in sys.argv
SKIP_PARTS    = "--skip-parts"    in sys.argv
SKIP_SUPPLIERS= "--skip-suppliers" in sys.argv

BASE_HEADERS = {
    "apikey":        SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type":  "application/json",
}

# ── HTTP helpers ──────────────────────────────────────────────────────────────

def _open(req):
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            raw = r.read()
            return json.loads(raw) if raw else []
    except urllib.error.HTTPError as e:
        body = e.read().decode()[:400]
        print(f"\n  [HTTP {e.code}] {body}")
        return None


def get_all(table, select="*"):
    results, offset, limit = [], 0, 1000
    while True:
        url = f"{SUPABASE_URL}/rest/v1/{table}?select={select}&limit={limit}&offset={offset}"
        req = urllib.request.Request(url, headers=BASE_HEADERS)
        batch = _open(req) or []
        results.extend(batch)
        if len(batch) < limit:
            break
        offset += limit
    return results


def batch_insert(table, rows, on_conflict=None):
    """Plain INSERT; optionally merge-duplicates on a named conflict column."""
    total, done = len(rows), 0
    prefer = "resolution=merge-duplicates,return=minimal" if on_conflict else "return=minimal"
    headers = {**BASE_HEADERS, "Prefer": prefer}
    for i in range(0, total, BATCH_SIZE):
        batch = rows[i : i + BATCH_SIZE]
        url = f"{SUPABASE_URL}/rest/v1/{table}"
        if on_conflict:
            url += f"?on_conflict={on_conflict}"
        req = urllib.request.Request(url, data=json.dumps(batch).encode(), headers=headers, method="POST")
        result = _open(req)
        if result is not None:
            done += len(batch)
        n = min(i + BATCH_SIZE, total)
        print(f"  {n}/{total}", end="\r", flush=True)
    print(f"  {done}/{total} rows written         ")
    return done


def insert_one(table, row):
    headers = {**BASE_HEADERS, "Prefer": "return=representation"}
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/{table}",
        data=json.dumps(row).encode(), headers=headers, method="POST"
    )
    result = _open(req)
    return result[0] if result else None

# ── Helpers ───────────────────────────────────────────────────────────────────

def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return None if s in ("", "-", "—") else s


def to_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None

# ── Step 1: parts ─────────────────────────────────────────────────────────────

def load_parts(ws):
    parts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku, desc, cost, new_cost, bin_loc, price_a, price_b, cat, make, model, year = row[:11]
        sku = clean(sku)
        if not sku:
            continue
        parts.append({
            "sku":          sku,
            "name":         clean(desc) or "",
            "cost_price":   to_float(new_cost) or to_float(cost),
            "price":        to_float(price_a),
            "bin_location": clean(bin_loc),
            "category":     clean(cat),
            "make":         clean(make),
            "model":        clean(model),
            "year_range":   clean(year),
        })
    return parts

# ── Step 2: suppliers + part_suppliers ───────────────────────────────────────

def load_supplier_items(ws):
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sup_item_id, supplier_name, item_id = row[:3]
        sup_item_id   = clean(str(sup_item_id)) if sup_item_id is not None else None
        supplier_name = clean(supplier_name)
        item_id       = clean(item_id)
        if supplier_name and item_id:
            items.append({
                "supplier_part_no": sup_item_id or "",
                "supplier_name":    supplier_name,
                "sku":              item_id,
            })
    return items


def ensure_suppliers(names_needed, existing_suppliers):
    name_to_id = {s["name"]: s["id"] for s in existing_suppliers}
    missing = [n for n in names_needed if n not in name_to_id]
    if missing:
        print(f"  Creating {len(missing)} new supplier(s): {', '.join(missing[:10])}{'…' if len(missing) > 10 else ''}")
        if not DRY_RUN:
            for name in missing:
                created = insert_one("suppliers", {"name": name})
                if created:
                    name_to_id[name] = created["id"]
                    print(f"    + {name} → id {created['id']}")
    return name_to_id

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"Loading {XLSX_PATH} …")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    # ── Parts ─────────────────────────────────────────────────────────────────
    if not SKIP_PARTS:
        print("\n── Step 1: parts ────────────────────────────────────────────")
        parts = load_parts(wb["Sheet1"])
        print(f"  {len(parts)} parts read from Sheet1")
        if DRY_RUN:
            print("  [dry-run] skipping insert")
        else:
            print("  Fetching existing SKUs …")
            existing = get_all("parts", "sku")
            existing_skus = {r["sku"] for r in existing if r.get("sku")}
            new_parts = [p for p in parts if p["sku"] not in existing_skus]
            print(f"  {len(existing_skus)} existing, {len(new_parts)} new to insert")
            batch_insert("parts", new_parts)
    else:
        print("\n── Step 1: parts [skipped] ──────────────────────────────────")

    # ── Supplier items ────────────────────────────────────────────────────────
    if not SKIP_SUPPLIERS:
        print("\n── Step 2: supplier items ───────────────────────────────────")
        sup_items = load_supplier_items(wb["SupplierItem"])
        print(f"  {len(sup_items)} supplier-item rows read from SupplierItem")

        unique_suppliers = sorted({r["supplier_name"] for r in sup_items})
        print(f"  {len(unique_suppliers)} unique suppliers: {', '.join(unique_suppliers[:15])}{'…' if len(unique_suppliers) > 15 else ''}")

        if not DRY_RUN:
            print("  Fetching existing suppliers …")
            existing_suppliers = get_all("suppliers", "id,name")
            name_to_id = ensure_suppliers(unique_suppliers, existing_suppliers)

            print("  Fetching all parts (sku → id) …")
            all_parts = get_all("parts", "id,sku")
            sku_to_id = {p["sku"]: p["id"] for p in all_parts}
            print(f"  {len(sku_to_id)} parts in DB")

            ps_rows, skipped = [], 0
            for r in sup_items:
                part_id     = sku_to_id.get(r["sku"])
                supplier_id = name_to_id.get(r["supplier_name"])
                if not part_id or not supplier_id:
                    skipped += 1
                    continue
                ps_rows.append({
                    "part_id":          part_id,
                    "supplier_id":      supplier_id,
                    "supplier_part_no": r["supplier_part_no"],
                })

            print(f"  {len(ps_rows)} part_supplier links (skipped {skipped} unmatched rows)")
            batch_insert("part_suppliers", ps_rows)
        else:
            print("  [dry-run] skipping upsert")
    else:
        print("\n── Step 2: supplier items [skipped] ─────────────────────────")

    print("\nDone.")


if __name__ == "__main__":
    main()
