#!/usr/bin/env python3
"""
Import vehicle models + auto-link parts → Supabase

Input file:  docs/vehicles_import.xlsx
  Required columns: code | make | model | year_from | year_end
  (year_end can be blank = current/ongoing model)

What it does:
  1. Inserts each row into the `vehicles` table (skips duplicates by code)
  2. For every new vehicle, finds all parts whose SKU starts with that code
     e.g. vehicle code "AD01A" links to parts "AD01A-001TH", "AD01A-002BK", …
  3. Creates `part_fitments` rows  (skips if the link already exists)

Usage:
  python import_vehicles.py [--dry-run] [--skip-fitments]

  --dry-run        Preview counts only, nothing written to Supabase
  --skip-fitments  Insert vehicles only, skip the fitment-linking step
"""

import openpyxl
import urllib.request
import urllib.error
import json
import sys
import os

if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SUPABASE_URL = "https://lskouiyvdngdzaquurhk.supabase.co"
SUPABASE_KEY = "sb_publishable_De4neqOoFn1wFyiVzaNT0A_HzPAE3YW"
XLSX_PATH    = os.path.join(os.path.dirname(__file__), "docs", "vehicles_import.xlsx")
BATCH_SIZE   = 500

DRY_RUN        = "--dry-run"        in sys.argv
SKIP_FITMENTS  = "--skip-fitments"  in sys.argv

BASE_HEADERS = {
    "apikey":        SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type":  "application/json",
}

# ── HTTP helpers ──────────────────────────────────────────────────────────────

def _open(req):
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            raw = r.read()
            return json.loads(raw) if raw else []
    except urllib.error.HTTPError as e:
        body = e.read().decode()[:400]
        print(f"\n  [HTTP {e.code}] {body}")
        return None


def get_all(table, select="*", extra=""):
    results, offset, limit = [], 0, 1000
    while True:
        sep = "&" if extra else ""
        url = f"{SUPABASE_URL}/rest/v1/{table}?select={select}{sep}{extra}&limit={limit}&offset={offset}"
        req = urllib.request.Request(url, headers=BASE_HEADERS)
        batch = _open(req) or []
        results.extend(batch)
        if len(batch) < limit:
            break
        offset += limit
    return results


def batch_upsert(table, rows, on_conflict):
    """Upsert rows; merge on conflict column."""
    total, done = len(rows), 0
    headers = {**BASE_HEADERS, "Prefer": "resolution=merge-duplicates,return=minimal"}
    for i in range(0, total, BATCH_SIZE):
        batch = rows[i : i + BATCH_SIZE]
        url = f"{SUPABASE_URL}/rest/v1/{table}?on_conflict={on_conflict}"
        req = urllib.request.Request(url, data=json.dumps(batch).encode(),
                                     headers=headers, method="POST")
        result = _open(req)
        if result is not None:
            done += len(batch)
        print(f"  {min(i+BATCH_SIZE, total)}/{total}", end="\r", flush=True)
    print(f"  {done}/{total} rows written         ")
    return done


def batch_insert(table, rows):
    """Plain insert, ignore duplicates."""
    total, done = len(rows), 0
    headers = {**BASE_HEADERS, "Prefer": "return=minimal"}
    for i in range(0, total, BATCH_SIZE):
        batch = rows[i : i + BATCH_SIZE]
        url = f"{SUPABASE_URL}/rest/v1/{table}"
        req = urllib.request.Request(url, data=json.dumps(batch).encode(),
                                     headers=headers, method="POST")
        result = _open(req)
        if result is not None:
            done += len(batch)
        print(f"  {min(i+BATCH_SIZE, total)}/{total}", end="\r", flush=True)
    print(f"  {done}/{total} rows written         ")
    return done

# ── Read Excel ────────────────────────────────────────────────────────────────

def read_xlsx():
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: file not found → {XLSX_PATH}")
        print("Save your spreadsheet as docs/vehicles_import.xlsx and re-run.")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active

    # Normalise header row
    headers = [str(c.value).strip().lower().replace(" ", "_") if c.value else ""
               for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Flexible column mapping
    COL_MAP = {
        "code":      ["code"],
        "make":      ["make"],
        "model":     ["model"],
        "year_from": ["year_from", "year from", "yearfrom", "from"],
        "year_end":  ["year_end",  "year end",  "yearend",  "to", "year_to"],
    }

    idx = {}
    for field, aliases in COL_MAP.items():
        for alias in aliases:
            if alias in headers:
                idx[field] = headers.index(alias)
                break
        if field not in idx:
            print(f"ERROR: could not find column '{field}' in spreadsheet.")
            print(f"  Found columns: {headers}")
            sys.exit(1)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[idx["code"]] or "").strip()
        if not code:
            continue
        make      = str(row[idx["make"]]  or "").strip()
        model     = str(row[idx["model"]] or "").strip()
        year_from = row[idx["year_from"]]
        year_end  = row[idx["year_end"]]

        try: year_from = int(year_from) if year_from else None
        except: year_from = None
        try: year_end = int(year_end) if year_end else None
        except: year_end = None

        rows.append({
            "code":      code,
            "make":      make,
            "model":     model,
            "year_from": year_from,
            "year_to":   year_end,   # vehicles table uses year_to
        })

    return rows

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  Vehicle Import + Fitment Linker")
    print("=" * 55)

    # 1. Read spreadsheet
    print(f"\n→ Reading {XLSX_PATH} …")
    vehicle_rows = read_xlsx()
    print(f"  {len(vehicle_rows)} vehicles found in spreadsheet")

    if DRY_RUN:
        print("\n[DRY RUN] No data written.")
        # Show sample
        for r in vehicle_rows[:5]:
            print(f"  {r}")
        if len(vehicle_rows) > 5:
            print(f"  … and {len(vehicle_rows)-5} more")
        return

    # 2. Check that the `code` column exists
    print("\n→ Checking vehicles table …")
    probe = get_all("vehicles", "id,code")
    if probe is None:
        print("\nERROR: Could not read the `code` column from the vehicles table.")
        print("  Run this SQL in Supabase first, then re-run this script:")
        print()
        print("    ALTER TABLE vehicles ADD COLUMN IF NOT EXISTS code TEXT;")
        print("    CREATE INDEX IF NOT EXISTS idx_vehicles_code ON vehicles(code);")
        print()
        sys.exit(1)

    existing_codes = {v["code"] for v in probe if v.get("code")}
    print(f"  {len(existing_codes)} vehicles with codes already in DB")

    new_vehicles = [v for v in vehicle_rows if v["code"] not in existing_codes]
    skipped = len(vehicle_rows) - len(new_vehicles)
    print(f"  {len(new_vehicles)} new  |  {skipped} already exist (will skip)")

    if new_vehicles:
        print(f"\n→ Inserting {len(new_vehicles)} vehicles …")
        batch_insert("vehicles", new_vehicles)

    if SKIP_FITMENTS:
        print("\n[--skip-fitments] Fitment linking skipped.")
        return

    # 3. Reload vehicles to get their IDs (including newly inserted)
    print("\n→ Reloading vehicles to get IDs …")
    all_vehicles = get_all("vehicles", "id,code")
    code_to_id = {v["code"]: v["id"] for v in all_vehicles if v.get("code")}
    print(f"  {len(code_to_id)} vehicles with codes loaded")

    if len(code_to_id) == 0:
        print("\nERROR: No vehicles have a code value — the `code` column may be empty.")
        print("  Make sure you ran the ALTER TABLE SQL and re-run this script.")
        sys.exit(1)

    # 4. Fetch all parts (id + sku only)
    print("\n→ Fetching parts …")
    parts = get_all("parts", "id,sku")
    print(f"  {len(parts)} parts loaded")

    # 5. Fetch existing fitments to avoid duplicates
    print("\n→ Fetching existing fitments …")
    existing_fitments = get_all("part_fitments", "part_id,vehicle_id")
    existing_pairs = {(str(f["part_id"]), str(f["vehicle_id"])) for f in existing_fitments}
    print(f"  {len(existing_pairs)} fitments already in DB")

    # 6. Build new fitments: part SKU starts with vehicle code + "-"
    print("\n→ Matching parts to vehicles by SKU prefix …")
    fitments_to_add = []
    match_counts = {}

    for part in parts:
        sku = (part.get("sku") or "").strip()
        if not sku:
            continue
        # Extract prefix: everything before the first "-"
        prefix = sku.split("-")[0].upper()
        if prefix in code_to_id:
            vid = code_to_id[prefix]
            pid = str(part["id"])
            if (pid, str(vid)) not in existing_pairs:
                fitments_to_add.append({"part_id": part["id"], "vehicle_id": vid})
                match_counts[prefix] = match_counts.get(prefix, 0) + 1

    print(f"\n  Matched {len(fitments_to_add)} new part→vehicle links across {len(match_counts)} vehicle codes")

    if match_counts:
        # Show top matches
        top = sorted(match_counts.items(), key=lambda x: -x[1])[:10]
        print("\n  Top matches by code:")
        for code, cnt in top:
            print(f"    {code}  →  {cnt} parts")

    if not fitments_to_add:
        print("\n  Nothing to link.")
        return

    print(f"\n→ Inserting {len(fitments_to_add)} fitments …")
    batch_insert("part_fitments", fitments_to_add)

    print("\n✅ Done!")
    print(f"   Vehicles inserted : {len(new_vehicles)}")
    print(f"   Fitments created  : {len(fitments_to_add)}")


if __name__ == "__main__":
    main()
