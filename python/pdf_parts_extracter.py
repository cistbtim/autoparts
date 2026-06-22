import argparse
import cv2
import numpy as np
import os
import re
import sys
import openpyxl

try:
    import pytesseract
    HAS_OCR = True
except ImportError:
    HAS_OCR = False


# ── Column detection ──────────────────────────────────────────────────────────

def detect_column_dividers(img):
    """
    Scan a full page image for vertical lines and return their x-positions
    as percentages of page width.
    """
    H, W = img.shape[:2]
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

    # Keep only vertical strokes spanning at least 25% of page height
    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, H // 4))
    v_mask   = cv2.morphologyEx(binary, cv2.MORPH_OPEN, v_kernel, iterations=1)

    # Sum each column — peaks are divider centres
    density = np.sum(v_mask > 0, axis=0).astype(float)
    if density.max() == 0:
        return []

    threshold = density.max() * 0.4
    dividers  = []
    in_peak   = False
    peak_xs   = []

    for x, d in enumerate(density):
        if d >= threshold:
            in_peak = True
            peak_xs.append(x)
        elif in_peak:
            dividers.append(int(np.mean(peak_xs)))
            peak_xs = []
            in_peak = False
    if in_peak and peak_xs:
        dividers.append(int(np.mean(peak_xs)))

    return [x / W * 100 for x in dividers]


def map_dividers_to_columns(dividers_pct):
    """
    Given detected divider percentages, assign them to catalogue column slots.
    Expected column order: [border | S/N | Picture | Our No. | Description | OEM | App | Unit | border]
    Returns dict of {key: (x0_pct, x1_pct)} or None if can't map.
    """
    d = sorted(dividers_pct)
    if len(d) < 5:
        return None   # not enough dividers found — fall back to user config

    # Drop near-zero and near-100 border lines, keep interior dividers
    interior = [x for x in d if 2 < x < 98]

    # We need at least 5 interior dividers to define 6 text columns
    # (S/N | pic_left | pic_right/partno_left | desc_left ... unit_right)
    if len(interior) < 4:
        return None

    # Heuristic: the picture column is the widest gap in the first third of the page
    # Typically: [sn_right, pic_left?, partno_right, desc_right, oem_right, app_right, unit_right]
    # But often S/N and picture share the left zone.
    # Strategy: take dividers, assign slots left-to-right.
    d = interior

    # If 4 dividers: pic_end, partno_end, desc_end, oem_end  (app+unit merged)
    # If 5 dividers: pic_end, partno_end, desc_end, oem_end, app_end
    # If 6+:         sn_end, pic_end, partno_end, desc_end, oem_end, app_end

    if len(d) >= 6:
        # Has explicit S/N divider before picture
        pic_x1     = d[1]
        partno_x1  = d[2]
        desc_x1    = d[3]
        oem_x1     = d[4]
        app_x1     = d[5]
        unit_x1    = d[6] if len(d) > 6 else 98.0
    elif len(d) == 5:
        pic_x1     = d[0]
        partno_x1  = d[1]
        desc_x1    = d[2]
        oem_x1     = d[3]
        app_x1     = d[4]
        unit_x1    = 98.0
    else:  # 4
        pic_x1     = d[0]
        partno_x1  = d[1]
        desc_x1    = d[2]
        oem_x1     = d[3]
        app_x1     = min(d[3] + (98 - d[3]) / 2, 95.0)
        unit_x1    = 98.0

    return {
        "pic":    (0.0,       pic_x1),
        "partno": (pic_x1,    partno_x1),
        "desc":   (partno_x1, desc_x1),
        "oem":    (desc_x1,   oem_x1),
        "app":    (oem_x1,    app_x1),
        "unit":   (app_x1,    unit_x1),
    }


# ── Row detection ─────────────────────────────────────────────────────────────

def detect_rows(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                   cv2.THRESH_BINARY_INV, 51, 15)
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
    h_lines  = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, h_kernel, iterations=2)
    cnts, _  = cv2.findContours(h_lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    W = img.shape[1]
    ys = sorted(cv2.boundingRect(c)[1] for c in cnts if cv2.boundingRect(c)[2] > W * 0.5)

    clean = []
    for y in ys:
        if not clean or y - clean[-1] > 20:
            clean.append(y)

    return [(clean[i], clean[i + 1]) for i in range(len(clean) - 1)
            if clean[i + 1] - clean[i] >= 40]


# ── OCR ───────────────────────────────────────────────────────────────────────

def ocr_row(row_img, col_bounds, W):
    """
    One tesseract call for the full text area of a row.
    Words are assigned to columns by their x-centre position.
    """
    result = {k: "" for k in col_bounds if k != "pic"}
    if not HAS_OCR or row_img.size == 0:
        return result

    text_keys  = [k for k in col_bounds if k != "pic"]
    roi_x0_pct = min(col_bounds[k][0] for k in text_keys)
    roi_x0     = int(W * roi_x0_pct / 100)
    roi        = row_img[:, roi_x0:]
    if roi.size == 0:
        return result

    h, w  = roi.shape[:2]
    scale = max(1, 80 // max(h, 1))
    if scale > 1:
        roi = cv2.resize(roi, (w * scale, h * scale), interpolation=cv2.INTER_CUBIC)

    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    try:
        data = pytesseract.image_to_data(thresh, config="--psm 6 --oem 3",
                                         output_type=pytesseract.Output.DICT)
    except Exception as e:
        print(f"    OCR error: {e}")
        return result

    cols = {k: [] for k in text_keys}
    for i, word in enumerate(data["text"]):
        word = word.strip()
        if not word or int(data["conf"][i]) < 20:
            continue
        word_x_px  = (data["left"][i] + data["width"][i] / 2) / scale + roi_x0
        word_x_pct = word_x_px / W * 100
        for key in text_keys:
            x0, x1 = col_bounds[key]
            if x0 <= word_x_pct < x1:
                cols[key].append(word)
                break

    return {k: " ".join(v) for k, v in cols.items()}


# ── Debug preview ─────────────────────────────────────────────────────────────

def save_preview(img, rows, col_bounds, out_path):
    """Draw detected rows and column boundaries on a copy of the image."""
    W  = img.shape[1]
    vis = img.copy()

    # Draw row bands
    for y0, y1 in rows:
        cv2.line(vis, (0, y0), (W, y0), (0, 200, 0), 1)
        cv2.line(vis, (0, y1), (W, y1), (0, 200, 0), 1)

    # Draw column dividers with labels
    colours = [(255, 80,  80),  (80,  80,  255), (80,  200, 80),
               (255, 200, 0),   (200, 0,   200), (0,   200, 200)]
    for i, (key, (x0, x1)) in enumerate(col_bounds.items()):
        px0 = int(W * x0 / 100)
        px1 = int(W * x1 / 100)
        col = colours[i % len(colours)]
        cv2.line(vis, (px0, 0), (px0, vis.shape[0]), col, 2)
        cv2.line(vis, (px1, 0), (px1, vis.shape[0]), col, 1)
        mid = (px0 + px1) // 2
        cv2.putText(vis, key, (mid - 20, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, col, 2)

    cv2.imwrite(out_path, vis)
    print(f"  Preview saved → {out_path}")


# ── Argument parsing ──────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--images",     required=True)
    p.add_argument("--out-images", required=True)
    p.add_argument("--out-excel",  required=True)
    p.add_argument("--pattern",    default=r"^[A-Z]{1,5}\d{3,8}[A-Z0-9]?$")
    p.add_argument("--tesseract",  default="")
    p.add_argument("--preview",    action="store_true",
                   help="Save a debug image showing detected rows and columns")

    # Manual column overrides (only used if auto-detection fails)
    p.add_argument("--pic-x0",        type=float, default=8.0)
    p.add_argument("--pic-x1",        type=float, default=21.0)
    p.add_argument("--col-partno-x0", type=float, default=21.0)
    p.add_argument("--col-partno-x1", type=float, default=32.0)
    p.add_argument("--col-desc-x0",   type=float, default=32.0)
    p.add_argument("--col-desc-x1",   type=float, default=65.0)
    p.add_argument("--col-oem-x0",    type=float, default=65.0)
    p.add_argument("--col-oem-x1",    type=float, default=78.0)
    p.add_argument("--col-app-x0",    type=float, default=78.0)
    p.add_argument("--col-app-x1",    type=float, default=91.0)
    p.add_argument("--col-unit-x0",   type=float, default=91.0)
    p.add_argument("--col-unit-x1",   type=float, default=98.0)
    return p.parse_args()


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    args = parse_args()
    a    = vars(args)

    IMAGE_FOLDER      = args.images
    OUTPUT_IMAGES_DIR = a["out_images"]
    OUTPUT_EXCEL_PATH = a["out_excel"]
    PART_PATTERN      = re.compile(args.pattern, re.IGNORECASE)

    if not os.path.isdir(IMAGE_FOLDER):
        print(f"ERROR: Image folder not found: {IMAGE_FOLDER}", file=sys.stderr)
        sys.exit(1)

    if HAS_OCR and args.tesseract:
        pytesseract.pytesseract.tesseract_cmd = args.tesseract

    os.makedirs(OUTPUT_IMAGES_DIR, exist_ok=True)
    supported   = (".png", ".jpg", ".jpeg", ".bmp")
    image_files = sorted(f for f in os.listdir(IMAGE_FOLDER) if f.lower().endswith(supported))

    if not image_files:
        print(f"ERROR: No images in {IMAGE_FOLDER}", file=sys.stderr)
        sys.exit(1)

    # ── Auto-detect column positions from first page ──────────────────────────
    first_img = cv2.imread(os.path.join(IMAGE_FOLDER, image_files[0]))
    col_bounds = None

    if first_img is not None:
        dividers = detect_column_dividers(first_img)
        col_bounds = map_dividers_to_columns(dividers)
        if col_bounds:
            print("Auto-detected column positions:")
            for k, (x0, x1) in col_bounds.items():
                print(f"  {k:10s}: {x0:.1f}% – {x1:.1f}%")
        else:
            print(f"Auto-detection found {len(dividers)} dividers — not enough, using manual config.")

    # Fall back to user-configured values
    if col_bounds is None:
        col_bounds = {
            "pic":    (a["pic_x0"],        a["pic_x1"]),
            "partno": (a["col_partno_x0"], a["col_partno_x1"]),
            "desc":   (a["col_desc_x0"],   a["col_desc_x1"]),
            "oem":    (a["col_oem_x0"],    a["col_oem_x1"]),
            "app":    (a["col_app_x0"],    a["col_app_x1"]),
            "unit":   (a["col_unit_x0"],   a["col_unit_x1"]),
        }
        print("Using manual column config:")
        for k, (x0, x1) in col_bounds.items():
            print(f"  {k:10s}: {x0:.1f}% – {x1:.1f}%")

    print(f"\nFound {len(image_files)} page images.\n")

    all_rows = []
    skipped  = 0
    preview_saved = False

    for page_idx, filename in enumerate(image_files):
        img = cv2.imread(os.path.join(IMAGE_FOLDER, filename))
        if img is None:
            print(f"  Skipping {filename}: unreadable.")
            continue

        W    = img.shape[1]
        rows = detect_rows(img)
        if not rows:
            print(f"  Page {page_idx+1} ({filename}): no rows detected.")
            continue

        # Save preview on first page with rows
        if args.preview and not preview_saved:
            preview_path = os.path.join(OUTPUT_IMAGES_DIR, "_preview_columns.png")
            save_preview(img, rows, col_bounds, preview_path)
            preview_saved = True

        print(f"  Page {page_idx+1} ({filename}): {len(rows)} rows — ", end="", flush=True)

        page_count = 0
        for row_idx, (y0, y1) in enumerate(rows):
            row_img = img[y0:y1, :]

            ocr = ocr_row(row_img, col_bounds, W)

            # Find part number in the Our No. column
            part_no = ""
            for word in ocr.get("partno", "").split():
                if PART_PATTERN.match(word.upper()):
                    part_no = word.upper()
                    break

            if HAS_OCR and not part_no:
                skipped += 1
                continue

            if not part_no:
                part_no = f"PART_{page_idx+1:03d}_{row_idx+1:03d}"

            # Crop and save picture column
            pic_x0   = int(W * col_bounds["pic"][0] / 100)
            pic_x1   = int(W * col_bounds["pic"][1] / 100)
            pic_crop = row_img[:, pic_x0:pic_x1]

            img_name = f"{part_no}.png"
            out_path = os.path.join(OUTPUT_IMAGES_DIR, img_name)
            if os.path.exists(out_path):
                img_name = f"{part_no}_p{page_idx+1}r{row_idx+1}.png"
                out_path = os.path.join(OUTPUT_IMAGES_DIR, img_name)
            cv2.imwrite(out_path, pic_crop)

            all_rows.append({
                "Source Page":                filename,
                "Our No.":                    part_no,
                "Description":                ocr.get("desc", ""),
                "OEM Reference":              ocr.get("oem",  ""),
                "Vehicle Application & Specs": ocr.get("app", ""),
                "Unit":                       ocr.get("unit", ""),
                "Saved Image Asset":          img_name,
            })
            page_count += 1

        print(f"{page_count} extracted.")

    print(f"\nTotal extracted: {len(all_rows)}  |  Skipped (headers/blanks): {skipped}")

    if not all_rows:
        print("\nNo parts extracted.")
        print("Tip: run with --preview to see a _preview_columns.png in your output folder.")
        print("     Then adjust column positions in the UI to match your catalogue.")
        return

    # ── Excel ──────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parts Catalogue Data"

    headers = ["Source Page", "Our No.", "Description", "OEM Reference",
               "Vehicle Application & Specs", "Unit", "Saved Image Asset"]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font      = openpyxl.styles.Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill      = openpyxl.styles.PatternFill(start_color="2B4C7E", end_color="2B4C7E", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(
            horizontal="center" if col_num in [1, 2, 6, 7] else "left", vertical="center")
    ws.row_dimensions[1].height = 28

    thin = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style="thin", color="D1D5DB"),
        right=openpyxl.styles.Side(style="thin", color="D1D5DB"),
        top=openpyxl.styles.Side(style="thin", color="D1D5DB"),
        bottom=openpyxl.styles.Side(style="thin", color="D1D5DB"),
    )

    for r_idx, row in enumerate(all_rows, start=2):
        vals = [row["Source Page"], row["Our No."], row["Description"],
                row["OEM Reference"], row["Vehicle Application & Specs"],
                row["Unit"], row["Saved Image Asset"]]
        ws.append(vals)
        ws.row_dimensions[r_idx].height = 45
        for c_idx in range(1, len(vals) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font   = openpyxl.styles.Font(name="Segoe UI", size=10)
            cell.border = thin
            cell.alignment = openpyxl.styles.Alignment(
                horizontal="center" if c_idx in [1, 2, 6, 7] else "left",
                vertical="center", wrap_text=(c_idx not in [1, 2, 6, 7]))

    for col in ws.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        max_len = max((len(str(c.value).split("\n")[0]) for c in col if c.value), default=8)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

    wb.save(OUTPUT_EXCEL_PATH)
    print(f"Images → {OUTPUT_IMAGES_DIR}")
    print(f"Excel  → {OUTPUT_EXCEL_PATH}")


if __name__ == "__main__":
    main()
