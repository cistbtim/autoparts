import os
import requests
import openpyxl

# ==============================================================================
# CONFIGURATION
# ==============================================================================
EXCEL_PATH    = r"C:\Users\Tim\Desktop\PDF_IMAGE_SAMPLE\new.xlsx"

SUPABASE_URL  = "https://lskouiyvdngdzaquurhk.supabase.co"
SUPABASE_KEY  = "sb_publishable_De4neqOoFn1wFyiVzaNT0A_HzPAE3YW"

# Column names in the Excel (must match exactly)
PART_NO_COL   = "Our No."
IMAGE_URL_COL = "Image URL"

# ==============================================================================

HEADERS = {
    "apikey":        SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type":  "application/json",
    "Prefer":        "return=minimal",
}


def update_image_url(part_no, image_url):
    url = f"{SUPABASE_URL}/rest/v1/supplier_catalogue?supplier_part_no=eq.{part_no}"
    r = requests.patch(url, headers=HEADERS, json={"image_url": image_url}, timeout=15)
    return r.status_code


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    if PART_NO_COL not in headers:
        print(f"❌  Column '{PART_NO_COL}' not found. Headers: {headers}")
        return
    if IMAGE_URL_COL not in headers:
        print(f"❌  Column '{IMAGE_URL_COL}' not found — run upload_to_gdrive.py first.")
        return

    part_col = headers.index(PART_NO_COL) + 1
    url_col  = headers.index(IMAGE_URL_COL) + 1

    updated = 0
    skipped = 0
    errors  = 0

    for row in ws.iter_rows(min_row=2):
        part_no   = str(row[part_col - 1].value or "").strip()
        image_url = str(row[url_col  - 1].value or "").strip()

        if not part_no or not image_url.startswith("http"):
            skipped += 1
            continue

        status = update_image_url(part_no, image_url)
        if status in (200, 204):
            updated += 1
            print(f"  ✓ {part_no}")
        else:
            errors += 1
            print(f"  ✗ {part_no}  HTTP {status}")

    print(f"\n✅  Done — {updated} updated, {skipped} skipped, {errors} errors.")


if __name__ == "__main__":
    main()
