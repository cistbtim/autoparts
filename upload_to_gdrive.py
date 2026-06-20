import os
import pickle
import openpyxl
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

# ==============================================================================
# CONFIGURATION
# ==============================================================================
IMAGES_FOLDER    = r"C:\Users\Tim\Desktop\PDF_IMAGE_SAMPLE\IMAGES"
EXCEL_PATH       = r"C:\Users\Tim\Desktop\PDF_IMAGE_SAMPLE\new.xlsx"
GDRIVE_FOLDER_ID = "1V8p-3JGiWKG80F0ccOLhQJKf7AvtqT_f"

# OAuth client secret JSON downloaded from Google Cloud Console
# (APIs & Services → Credentials → OAuth 2.0 Client IDs → Desktop app → Download JSON)
OAUTH_CLIENT_FILE = r"C:\Users\Tim\Desktop\oauth_client.json"

# Token is saved here after first login so you don't need to re-authorise each run
TOKEN_FILE = r"C:\Users\Tim\Desktop\gdrive_token.pkl"

SCOPES = ["https://www.googleapis.com/auth/drive.file"]

# Column names in the Excel
PART_NO_COLUMN   = "Our No."
IMAGE_URL_COLUMN = "Image URL"

# ==============================================================================


def get_drive_service():
    creds = None
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, "rb") as f:
            creds = pickle.load(f)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(OAUTH_CLIENT_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "wb") as f:
            pickle.dump(creds, f)

    return build("drive", "v3", credentials=creds, cache_discovery=False)


def upload_file(drive, local_path, filename):
    """Upload one PNG to the Drive folder, make it public, return the view URL."""
    meta  = {"name": filename, "parents": [GDRIVE_FOLDER_ID]}
    media = MediaFileUpload(local_path, mimetype="image/png", resumable=False)
    f = drive.files().create(body=meta, media_body=media, fields="id").execute()
    file_id = f["id"]
    drive.permissions().create(
        fileId=file_id,
        body={"role": "reader", "type": "anyone"},
    ).execute()
    return f"https://drive.google.com/thumbnail?id={file_id}&sz=w200"


def main():
    if not os.path.exists(OAUTH_CLIENT_FILE):
        print(f"❌  OAuth client file not found: {OAUTH_CLIENT_FILE}")
        print()
        print("Steps to create it:")
        print("  1. Go to console.cloud.google.com → APIs & Services → Credentials")
        print("  2. Click '+ Create Credentials' → OAuth 2.0 Client ID")
        print("  3. Application type: Desktop app → Create")
        print("  4. Download the JSON → save as C:\\Users\\Tim\\Desktop\\oauth_client.json")
        return

    # Collect all PNG files
    png_files = {
        os.path.splitext(f)[0]: os.path.join(IMAGES_FOLDER, f)
        for f in os.listdir(IMAGES_FOLDER)
        if f.lower().endswith(".png")
    }
    print(f"Found {len(png_files)} images in {IMAGES_FOLDER}")

    # Load the Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if PART_NO_COLUMN not in headers:
        print(f"❌  Column '{PART_NO_COLUMN}' not found. Headers: {headers}")
        return

    part_no_col = headers.index(PART_NO_COLUMN) + 1

    if IMAGE_URL_COLUMN in headers:
        url_col = headers.index(IMAGE_URL_COLUMN) + 1
    else:
        url_col = len(headers) + 1
        from openpyxl.styles import Font, PatternFill, Alignment
        h = ws.cell(row=1, column=url_col)
        h.value     = IMAGE_URL_COLUMN
        h.font      = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        h.fill      = PatternFill(start_color="2B4C7E", end_color="2B4C7E", fill_type="solid")
        h.alignment = Alignment(horizontal="center", vertical="center")
        print(f"Added '{IMAGE_URL_COLUMN}' column at position {url_col}")

    # Authorise and connect (opens browser on first run)
    print("Connecting to Google Drive (browser will open if first run)…")
    drive = get_drive_service()
    print("Connected ✓")

    uploaded = 0
    skipped  = 0

    for row in ws.iter_rows(min_row=2):
        part_no = str(row[part_no_col - 1].value or "").strip()
        if not part_no:
            continue

        existing_url = str(row[url_col - 1].value or "").strip()
        if existing_url.startswith("http"):
            skipped += 1
            continue

        local_path = png_files.get(part_no)
        if not local_path:
            print(f"  ⚠  No image file for {part_no}")
            continue

        try:
            url = upload_file(drive, local_path, f"{part_no}.png")
            row[url_col - 1].value = url
            uploaded += 1
            print(f"  ✓ {part_no:15s} → {url}")
        except Exception as e:
            print(f"  ✗ {part_no}: {e}")

    wb.save(EXCEL_PATH)
    print(f"\n✅  Done — {uploaded} uploaded, {skipped} already had URLs.")
    print(f"    Excel updated → {EXCEL_PATH}")


if __name__ == "__main__":
    main()
